"""
export.py

Owns Excel workbook assembly and file export.
Receives the clean DataFrame, analysis results dict, chart image
paths, run_id, and output directory. Assembles a six-sheet .xlsx
workbook and returns the output file path.

Side effect: writes one .xlsx file to output/.
"""

from pathlib import Path
from datetime import date
import csv

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment

import config


# Module-level — not logic, just a lookup table
CLEAN_DATA_HEADERS = {
    config.COL_DATETIME:       "Timestamp",
    config.COL_ACTIVE_POWER:   "Active Power (kW)",
    config.COL_WIND_SPEED:     "Wind Speed (m/s)",
    config.COL_THEORETICAL:    "Theoretical Power (kWh)",
    config.COL_WIND_DIRECTION: "Wind Direction (°)",
}

EFFICIENCY_HEADERS = {
    config.COL_DATETIME:       "Timestamp",
    config.COL_ACTIVE_POWER:   "Active Power (kW)",
    config.COL_WIND_SPEED:     "Wind Speed (m/s)",
    config.COL_THEORETICAL:    "Theoretical Power (kWh)",
    "efficiency_ratio":        "Efficiency Ratio",
}


def export(
    clean_df: pd.DataFrame,
    results: dict[str, pd.DataFrame],
    chart_paths: list[Path],
    run_id: str,
    output_dir: Path,
) -> Path:
    """
    Assemble and save the six-sheet Excel report.

    Parameters:
        clean_df (pd.DataFrame): cleaned SCADA DataFrame from clean.clean()
        results (dict[str, pd.DataFrame]): analysis results from
            analyse.analyse(). Must contain keys: 'stats', 'efficiency',
            'monthly', 'daily', 'wind_bins'.
        chart_paths (list[Path]): list of three .png paths from
            visualise.visualise(), in order:
            [power_trend.png, wind_scatter.png, monthly_bar.png]
        run_id (str): UUID from clean.clean(), used to filter the
            quality log to this run's entries only
        output_dir (Path): directory to write the report to

    Returns:
        Path: full path of the written .xlsx file,
              e.g. output/report_2026-05-20.xlsx

    Assumptions:
        - output_dir exists
        - All three chart .png files in chart_paths exist on disk
        - logs/data_quality.log exists and contains entries for run_id
    """
    wb = Workbook()
    wb.remove(wb.active)  # Remove the default empty sheet

    # Add sheets in required order
    _write_summary(wb.create_sheet("Summary"), results, clean_df)
    _write_dataframe(
        wb.create_sheet("Clean Data"),
        clean_df,
        plain_headers=CLEAN_DATA_HEADERS,
    )
    _write_trend_analysis(wb.create_sheet("Trend Analysis"), results)
    _write_power_curve(wb.create_sheet("Power Curve Analysis"), results["efficiency"])
    _write_charts(wb.create_sheet("Charts"), chart_paths)
    _write_quality_log(wb.create_sheet("Data Quality Log"), run_id)

    # Build timestamped filename — Invariant 6
    base_filename = f"report_{date.today().isoformat()}"
    filename = f"{base_filename}.xlsx"
    out_path = output_dir / filename
    
    counter = 1
    while out_path.exists():
        filename = f"{base_filename}_{counter}.xlsx"
        out_path = output_dir / filename
        counter += 1
        
    wb.save(out_path)

    return out_path


def _write_summary(
    ws, results: dict, clean_df: pd.DataFrame
) -> None:
    """
    Write the Summary sheet with narrative paragraph and statistics table.

    Component 1 (rows 1-3): Narrative paragraph merged across A1:F3.
    Component 2 (rows 5+): Headline statistics table with plain-English
    labels and styled header.

    Parameters:
        ws: openpyxl Worksheet
        results (dict): analysis results from analyse.analyse()
        clean_df (pd.DataFrame): cleaned SCADA DataFrame

    Returns:
        None
    """
    # ── Component 1: Narrative paragraph ────────────────────────────
    stats = results["stats"]
    efficiency_df = results["efficiency"]
    monthly_df = results["monthly"]

    peak_month = monthly_df[config.COL_ACTIVE_POWER].idxmax()
    peak_month_label = peak_month.strftime("%B %Y")
    mean_power = stats.loc["mean", config.COL_ACTIVE_POWER]
    mean_efficiency = efficiency_df["efficiency_ratio"].mean() * 100
    op_hours = round(len(efficiency_df) * 10 / 60)

    narrative = (
        f"This report analyses {op_hours:,} hours of operational wind turbine data. "
        f"The turbine produced a mean active power output of {mean_power:,.1f} kW "
        f"across all operational periods. "
        f"Overall efficiency against the theoretical power curve averaged "
        f"{mean_efficiency:.1f}%, indicating the proportion of available wind energy "
        f"converted to output. "
        f"Peak mean output occurred in {peak_month_label}, suggesting strong seasonal "
        f"wind resource during this period."
    )

    # Write narrative into merged cells A1:F3 with wrap_text
    ws.merge_cells("A1:F3")
    cell = ws.cell(row=1, column=1, value=narrative)
    cell.alignment = Alignment(wrap_text=True, vertical="top")

    # ── Component 2: Headline statistics table (rows 5 onward) ──────
    stats_table = [
        ("Metric", "Value"),
        ("Mean Active Power (kW)",
         round(stats.loc["mean", config.COL_ACTIVE_POWER], 2)),
        ("Max Active Power (kW)",
         round(stats.loc["max", config.COL_ACTIVE_POWER], 2)),
        ("Std Dev Active Power (kW)",
         round(stats.loc["std", config.COL_ACTIVE_POWER], 2)),
        ("Mean Wind Speed (m/s)",
         round(stats.loc["mean", config.COL_WIND_SPEED], 2)),
        ("Max Wind Speed (m/s)",
         round(stats.loc["max", config.COL_WIND_SPEED], 2)),
        ("Mean Efficiency Ratio (%)",
         round(efficiency_df["efficiency_ratio"].mean() * 100, 2)),
        ("Total Rows Analysed", len(efficiency_df)),
        ("Rows Excluded from Efficiency",
         len(clean_df) - len(efficiency_df)),
    ]

    for row_idx, (label, value) in enumerate(stats_table, start=5):
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=value)

    # Style the header row (row 5)
    _style_header_row(ws, row_num=5, col_count=2)


def _write_dataframe(
    ws,
    df: pd.DataFrame,
    plain_headers: dict,
) -> None:
    """
    Write a DataFrame to a worksheet with plain-English headers.

    Renames columns using the plain_headers mapping before writing.
    Styles row 1 as a header row (dark fill, white bold text).
    All other rows are written as plain data.

    Parameters:
        ws: openpyxl Worksheet to write to
        df (pd.DataFrame): DataFrame to write
        plain_headers (dict): mapping of raw column names to
                              plain-English display labels.
                              Columns not in the mapping are
                              written with their original names.

    Returns:
        None
    """
    display_df = df.rename(columns=plain_headers)

    for row_idx, row in enumerate(
        dataframe_to_rows(display_df, index=True, header=True), start=1
    ):
        ws.append(row)
        if row_idx == 1:
            _style_header_row(ws, row_num=1, col_count=len(row))


def _write_trend_analysis(ws, results: dict) -> None:
    """
    Write the Trend Analysis sheet with monthly and daily sections.

    Layout:
      - Row 1: bold label "Monthly Mean Active Power"
      - Row 2: header row
      - Rows 3-N: monthly data
      - One blank row
      - Bold label "Daily Total Active Power"
      - Header row
      - Remaining rows: daily data

    Parameters:
        ws: openpyxl Worksheet
        results (dict): analysis results from analyse.analyse()

    Returns:
        None
    """
    monthly_df = results["monthly"]
    daily_df = results["daily"]

    # ── Monthly section ────────────────────────────────────────────
    ws.cell(row=1, column=1, value="Monthly Mean Active Power").font = Font(
        bold=True, size=11
    )

    header_row = 2
    monthly_headers = ["Date", "Active Power (kW)"]
    ws.append(monthly_headers)
    _style_header_row(ws, row_num=header_row, col_count=len(monthly_headers))

    for idx, (date_val, power) in enumerate(monthly_df.iterrows()):
        ws.append([str(date_val.date()), round(power[config.COL_ACTIVE_POWER], 2)])

    # ── Daily section ──────────────────────────────────────────────
    # Position after monthly data + 1 blank row + 1 label row
    daily_start_row = len(monthly_df) + 3 + 1  # monthly rows + blank + header

    ws.cell(row=daily_start_row, column=1, value="Daily Total Active Power").font = (
        Font(bold=True, size=11)
    )

    daily_header_row = daily_start_row + 1
    daily_headers = ["Date", "Active Power (kW)"]
    # Write header using ws.cell to avoid appending after previous append offset
    for col_idx, header in enumerate(daily_headers, start=1):
        ws.cell(row=daily_header_row, column=col_idx, value=header)
    _style_header_row(ws, row_num=daily_header_row, col_count=len(daily_headers))

    for row_offset, (date_val, power) in enumerate(daily_df.iterrows(), start=1):
        row_num = daily_header_row + row_offset
        ws.cell(row=row_num, column=1, value=str(date_val.date()))
        ws.cell(row=row_num, column=2, value=round(power[config.COL_ACTIVE_POWER], 2))


def _write_power_curve(ws, efficiency_df: pd.DataFrame) -> None:
    """
    Write the efficiency DataFrame to the Power Curve Analysis sheet.

    Parameters:
        ws: openpyxl Worksheet
        efficiency_df (pd.DataFrame): efficiency results from
                                      analyse._compute_efficiency()

    Returns:
        None
    """
    _write_dataframe(ws, efficiency_df, plain_headers=EFFICIENCY_HEADERS)


def _write_charts(ws, chart_paths: list[Path]) -> None:
    """
    Embed chart .png files as images into the Charts sheet.

    Images are not linked — they are embedded directly into
    the workbook so the file is self-contained.

    Parameters:
        ws: openpyxl Worksheet
        chart_paths (list[Path]): list of three .png file paths,
                                  in order: power_trend, wind_scatter,
                                  monthly_bar

    Returns:
        None

    Assumptions:
        - All paths in chart_paths exist on disk
    """
    # Anchor positions — stacked vertically with spacing
    anchors = ["A1", "A32", "A62"]
    titles = [
        "Daily Power Trend",
        "Wind Speed vs Power Output",
        "Monthly Mean Power",
    ]

    for path, anchor, title in zip(chart_paths, anchors, titles):
        # Write a plain label above each image
        row_num = int(anchor[1:]) if len(anchor) > 2 else int(anchor[1])
        ws.cell(row=row_num, column=1, value=title).font = Font(
            bold=True, size=11
        )

        img = XLImage(str(path))
        img.width = 700
        img.height = 300
        # Anchor image two rows below the label
        img_anchor = f"A{row_num + 2}"
        ws.add_image(img, img_anchor)


def _write_quality_log(ws, run_id: str) -> None:
    """
    Write quality log entries for the current run to the sheet.

    Reads logs/data_quality.log, filters to the current run_id,
    and writes matching rows with a styled header.

    Parameters:
        ws: openpyxl Worksheet
        run_id (str): UUID identifying the current run

    Returns:
        None

    Assumptions:
        - config.LOG_PATH exists and is readable
        - At least one entry for run_id exists in the log
    """
    log_path: Path = config.LOG_PATH

    with open(log_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        rows = [r for r in reader if r[config.LOG_FIELD_RUN_ID] == run_id]

    # Write header
    headers = config.LOG_FIELDS
    ws.append(headers)
    _style_header_row(ws, row_num=1, col_count=len(headers))

    # Write matching rows
    for row in rows:
        ws.append([row[field] for field in headers])


def _style_header_row(ws, row_num: int, col_count: int) -> None:
    """
    Apply header styling to a row in a worksheet.

    Style: dark navy background (#1B3A5C), white bold text,
    centre-aligned.

    Parameters:
        ws: openpyxl Worksheet
        row_num (int): 1-based row number to style
        col_count (int): number of columns to style (starting from col 1)

    Returns:
        None
    """
    header_fill = PatternFill(
        start_color="1B3A5C", end_color="1B3A5C", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment